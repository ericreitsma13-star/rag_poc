from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from app.utils.config import SUPPORTED_EXTENSIONS


@dataclass(frozen=True)
class ParsedText:
    text: str
    metadata: dict


def _read_pdf(path: Path) -> list[ParsedText]:
    import fitz

    chunks: list[ParsedText] = []
    with fitz.open(path) as doc:
        for page_index, page in enumerate(doc):
            text = page.get_text("text").strip()
            if text:
                chunks.append(ParsedText(text=text, metadata={"page": page_index + 1}))
    return chunks


def _read_docx(path: Path) -> list[ParsedText]:
    import docx

    doc = docx.Document(path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip()).strip()
    return [ParsedText(text=text, metadata={})] if text else []


def _read_text(path: Path) -> list[ParsedText]:
    text = path.read_text(encoding="utf-8", errors="ignore").strip()
    return [ParsedText(text=text, metadata={})] if text else []


def _read_xlsx(path: Path, rows_per_chunk: int = 50) -> list[ParsedText]:
    """
    Parse an XLSX file into text chunks.

    Each chunk covers up to `rows_per_chunk` rows from one sheet.
    Metadata includes the sheet name and the row range (1-based, header excluded).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chunks: list[ParsedText] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = [str(c) if c is not None else "" for c in rows[0]]
        data_rows = rows[1:]

        for start in range(0, len(data_rows), rows_per_chunk):
            batch = data_rows[start : start + rows_per_chunk]
            lines: list[str] = []
            for row in batch:
                cells = [str(c) if c is not None else "" for c in row]
                line = " | ".join(f"{h}: {v}" for h, v in zip(header, cells) if h or v)
                if line.strip():
                    lines.append(line)

            if not lines:
                continue

            row_start = start + 2  # 1-based, skip header
            row_end = start + len(batch) + 1
            chunks.append(
                ParsedText(
                    text="\n".join(lines),
                    metadata={
                        "sheet": sheet_name,
                        "row_start": row_start,
                        "row_end": row_end,
                    },
                )
            )

    wb.close()
    return chunks


def parse_file(path: Path) -> list[ParsedText]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        return []

    if suffix == ".pdf":
        return _read_pdf(path)
    if suffix == ".docx":
        return _read_docx(path)
    if suffix == ".xlsx":
        return _read_xlsx(path)
    return _read_text(path)
